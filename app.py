import tkinter
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox

import tabula
import csv
import os
from openpyxl import Workbook, load_workbook
import platform
class PDF_Main_Setting:
    def __init__(self, import_file, export_file):
        self.import_file = import_file
        self.pdf_array = []
        self.export_file = export_file
        self.page_didnt_section = []
        self.os = platform.platform()

    def upload_data_print(self, marka, product, price):
            minta = {
                "Marka" : marka,
                "Product" : product,
                "Price" : price
            }
            self.pdf_array.append(minta)

    def phone_read(self, line):
            max_price = 0
            for column in line:

                if column == "Apple" or column == "Samsung":
                    var1 = column
                if "iPhone" in column or "GALAXY" in column or "APPLE WATCH" in column or "SAMSUNG GEAR" in column or "SAMSUNG WATCH" in column or "16GB" in column or "128GB" in column or "64GB" in column or "256GB" in column or "32GB" in column:
                    var2 = column
                if ("€" in column or '#N/D' == column) and max_price < 1:
                    max_price += 1
                    var3 = column
            self.upload_data_print(var1, var2.replace("  ", " "), var3)

    def phone_read_with_split(self, line):
            for column in line:
                if "Apple" in column or "Samsung" in column:
                    var1 = column.split(' ')[0]
                    price = column.split(' ')[len(column.split(' ')) - 2]
                    currency = column.split(' ')[len(column.split(' ')) - 1]
                    var3 = f"{price} {currency}"
                    var2 = column.replace(var1, "").replace(var3, "").strip()
                    self.upload_data_print(var1, var2.replace("  ", " "), var3)

    def computer_read(self, reader):
            product_list = []
            price_list = []
            var = ""
            j = 0
            for line in reader:
                max_price = 0

                for column in line:
                    if "MacBook" in column or "iMac" in column:
                        if var != "":
                            # test_function("Apple", var, var3)
                            # product_list.append(var)
                            j += 1
                        var = column
                        product_list.append(var)
                    elif "128GB" in column or "Space Gray" in column or "1TB" in column or "512GB" in column or "16GB" in column or "i5" in column or "i7" in column or "(2GB)" in column:
                        product_list[j] = f'{product_list[j]} {column}'
                    if ("€" in column or '#N/D' == column) and max_price < 1:
                        max_price += 1
                        price_list.append(column)
            j = 0
            while j < len(product_list):
                self.upload_data_print("Apple", product_list[j].replace("Apple ", ""), price_list[j])
                j += 1

    def csv_file_read(self, file):
            boolean_Test = False
            with open(file, 'r') as f:
                reader = csv.reader(f)

                for line in reader:
                    for column in line:
                        if "MacBook" in column or "iMac" in column:
                            boolean_Test = True
                            break

            with open(file, 'r') as f:
                reader = csv.reader(f)
                if boolean_Test:
                    self.computer_read(reader)
                else:
                    try:
                        for line in reader:
                            for column in line:
                                if "Apple" in column or "Samsung" in column:
                            #phone_read_with_split(line)
                                    self.phone_read(line)
                    except Exception as e:
                        for line in reader:
                            self.phone_read_with_split(line)
                            #phone_read(line)

    def read_pdf(self):
        if "Windows" in self.os:
            os.system("chcp 65001")


        dfs = tabula.read_pdf(self.import_file, pages='all')

            #print(dfs[1].to_csv("test.csv"))
        for i in range(len(dfs)):
            dfs[i].to_csv(f"test{i}.csv")
        for num in range(0, len(dfs)):
            self.csv_file_read(f"test{num}.csv")
            os.remove(f"test{num}.csv")


        wb = load_workbook(self.export_file)
        ws = wb.active
        #print(ws["B170"].value)

        for item in self.pdf_array:
                i = 1
                veg = True
                while veg:
                    try:
                        if ws[f"B{i}"].value.strip() == item['Product']:
                            #print(item['Product'])
                            ws[f"E{i}"].value = item['Price'].replace(" €", "")
                            try:
                                ws[f"F{i}"].value = float(item['Price'].replace(" €", "")) * 1.27
                            except ValueError:
                                ws[f"F{i}"].value = ws[f"E{i}"].value
                            #print(f'{item['Product']} {ws[f"E{i}"].value}')
                            wb.save(self.export_file)
                            veg = False
                        i += 1
                    except AttributeError:
                        if 'Apple Apple' not in item['Price'] or item['Product'] != "" or "IPAD" not in item['Product']:
                            self.page_didnt_section.append(item)
                        veg = False
        return self.page_didnt_section #Visszaadja azokat az ertekeket amik nem kerultek bele

    def print_data_section(self):
        wb = load_workbook(self.export_file)
        ws = wb.active



        if self.page_didnt_section != []:
            hossz = len(ws['A'])
            i = 0
            #print(self.page_didnt_section)
            while i < len(self.page_didnt_section):
                if str(self.page_didnt_section[i]['Product']) != "":
                    ws[f'A{hossz + i + 1}'].value = str(self.page_didnt_section[i]['Product']).replace(" ", "_").upper()
                    #wb.save(self.export_file)
                    ws[f'B{hossz + i + 1}'].value = self.page_didnt_section[i]['Product']
                    #wb.save(self.export_file)
                    ws[f'D{hossz + i + 1}'].value = 1
                    #wb.save(self.export_file)
                    ws[f'E{hossz + i + 1}'].value = str(self.page_didnt_section[i]['Price']).replace(" €", "")
                    #wb.save(self.export_file)
                    try:
                        ws[f"F{hossz + i + 1}"].value = float(self.page_didnt_section[i]['Price'].replace(" €", "")) * 1.27
                    except ValueError:
                        ws[f"F{hossz + i + 1}"].value = "invalid"
                    #wb.save(self.export_file)
                    ws[f'G{hossz + i + 1}'].value = 0
                    #wb.save(self.export_file)
                    ws[f'H{hossz + i + 1}'].value = 0
                    #wb.save(self.export_file)
                    if "WATCH" in self.page_didnt_section[i]['Product']:
                        ws[f'K{hossz + i + 1}'].value = f'Refurbished {"Samsung" if self.page_didnt_section[i]['Marka'] == "Samsung" else "Apple"} Smartwatch'
                    elif "iPhone" in self.page_didnt_section[i]['Product'] or "GALAXY" in self.page_didnt_section[i]["Product"]:
                        ws[f'K{hossz + i + 1}'].value = f'Refurbished {"Samsung" if self.page_didnt_section[i]['Marka'] == "Samsung" else "Apple"} Smartphones'
                    elif "IPAD" in self.page_didnt_section[i]['Product']:
                        ws[f'K{hossz + i + 1}'].value = "Refurbished IPAD"
                    elif "iMac" in self.page_didnt_section[i]['Product']:
                        ws[f'K{hossz + i + 1}'].value = "Refurbished IMAC"
                    elif "MacBook" in self.page_didnt_section[i]["Product"]:
                        ws[f'K{hossz + i + 1}'].value = "Refurbished MacBook"
                    #wb.save(self.export_file)
                    ws[f'L{hossz + i + 1}'].value = self.page_didnt_section[i]['Marka']
                    #wb.save(self.export_file)
                    ws[f'M{hossz + i + 1}'].value = "Für weitere Informationen klicken Sie bitte auf den Link unten:   ↘"
                    #wb.save(self.export_file)
                    ws[f'O{hossz + i + 1}'].value = f"https://mobil-tel.unas.hu/{str(self.page_didnt_section[i]['Product']).replace(" ", "_").upper()}"
                    #wb.save(self.export_file)
                    ws[f'P{hossz + i + 1}'].value = 1
                    #wb.save(self.export_file)
                    ws[f'R{hossz + i + 1}'].value = "db"
                    #wb.save(self.export_file)
                    ws[f'S{hossz + i + 1}'].value = "Colour:White/Weiss(+)0|Red/Rot(+)0|Yellow/Gelb(+)0|Green/Grün(+)0|Blue/Blau(+)0|Black/Schwarz(+)0"
                    #wb.save(self.export_file)
                    ws[f'T{hossz + i + 1}'].value = "GARANTIE:GARANTIE 3 MONATE(+)0|GARANTIE 6 MONATE(+)10|GARANTIE 12 MONATE(+)20"
                    wb.save(self.export_file)
                i+=1


class myGUI:
    def __init__(self):
        self.import_file = ""
        self.export_file = ""

        self.window = Tk()
        self.window.geometry("800x400")
        #self.window.iconbitmap("pdfscanner.ico")
        self.window.title("PDFScanner to XLSX write")


        self.label = Label(text="PDF Scanner", font=("Arial", 40), pady=50)
        self.label.pack()
        self.import_maintext = Label(text="Import")
        self.import_maintext.pack()
        self.import_text = Label(text=f'{self.import_file}')
        self.import_text.pack()
        self.import_button = Button(text="Import PDF File", command=self.openFile_import)
        self.import_button.pack()
        self.export_maintext = Label(text="Export")
        self.export_maintext.pack()
        self.export_Text = Label(text=self.export_file)
        self.export_Text.pack()
        self.export_button2 = Button(text="Export XLSX File", command=self.openFile_export)
        self.export_button2.pack()
        self.start_button = Button(text="Start", pady=10, padx=10, command=self.start_file)
        self.start_button.pack()

        self.window.mainloop()

    def openFile_import(self):
        filepath = filedialog.askopenfilename(
                                                title="Open file okay?",
                                              filetypes= (("PDF files","*.pdf"),
                                              ("all files","*.*")))
        print(filepath)
        file = open(filepath,'r')
        self.import_file = filepath
        file.close()
        return filepath

    def openFile_export(self):
        filepath = filedialog.askopenfilename(title="Open file okay?",
                                              filetypes= (("Excel files","*.xlsx"),
                                              ("all files","*.*")))
        print(filepath)
        file = open(filepath,'r')
        self.export_file = filepath
        file.close()

    def start_file(self):
        PDF = PDF_Main_Setting(self.import_file, self.export_file)
        termek = PDF.read_pdf()
        szoveg = ""
        for item in termek:
            if item["Product"] != "":
                szoveg += f'{item["Product"]}\n'
        #print(szoveg)
        if messagebox.askokcancel("Scan PDF", "Do you want to scan?"):
            PDF.print_data_section()
            messagebox.showinfo("Scan PDF", "Finish scan")

myGUI()
